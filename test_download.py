import time
import os
import requests
from selene import browser, query
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
# for pdf
from pypdf import PdfReader
# for xlsx
from openpyxl import load_workbook
# for xls
from xlrd import open_workbook
# for zip
from zipfile import ZipFile

def test_rst():
    current_file = os.path.abspath(__file__)
    current_dir = os.path.dirname(current_file)
    print(current_dir)
    print(current_file)
    file_dir = os.path.join(current_dir, "here")
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": file_dir,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    browser.config.driver = driver

    browser.open('https://github.com/pytest-dev/pytest/blob/main/README.rst')
    # browser.element('[data-testid=download-raw-button]').click()
    # time.sleep(7)
    download_url = browser.element('[data-testid=raw-button]').get(query.attribute('href'))
    print(download_url)
    content = requests.get(url=download_url).content
    with open("here/readme.rst", 'wb') as file:
        file.write(content)
    with open("here/readme.rst", 'r') as file:
        a = file.read()
        assert "The maintainers of pytest" in a
def test_pdf():
    current_file = os.path.abspath(__file__)
    current_dir = os.path.dirname(current_file)
    print(current_dir)
    print(current_file)
    file_dir = os.path.join(current_dir, "pdf")
    file_dir2= os.path.join(file_dir, "Python Testing with Pytest (Brian Okken).pdf")
    reader = PdfReader(file_dir2)
    print(reader.pages)
    print(len(reader.pages))
    print(reader.pages[1].extract_text())
    assert "Starter Kit" in reader.pages[1].extract_text()
    print(os.path.getsize(file_dir2))
    assert os.path.getsize(file_dir2) == 3035139

def test_xlsx():
    current_file = os.path.abspath(__file__)
    current_dir = os.path.dirname(current_file)
    print(current_dir)
    print(current_file)
    file_dir = os.path.join(current_dir, "xlsx")
    file_dir2 = os.path.join(file_dir, "file_example_XLSX_50.xlsx")
    workbook = load_workbook(file_dir2)
    sheet = workbook.active
    print(sheet.cell(row=1, column=2).value)
def test_xls():
    current_file = os.path.abspath(__file__)
    current_dir = os.path.dirname(current_file)
    print(current_dir)
    print(current_file)
    file_dir = os.path.join(current_dir, "xls")
    file_dir2 = os.path.join(file_dir, "file_example_XLS_10.xls")
    workbook = open_workbook(file_dir2)
    print(workbook.nsheets)
    print(workbook.sheet_names())
    sheet= workbook.sheet_by_index(0)
    print(sheet.nrows)
    print(sheet.ncols)
    print(sheet.cell_value(rowx=9, colx=3))
    for rx in range(sheet.nrows):
        print(sheet.row(rx))

def test_zip():
    current_file = os.path.abspath(__file__)
    current_dir = os.path.dirname(current_file)
    print(current_dir)
    print(current_file)
    file_dir = os.path.join(current_dir, "zip")
    file_dir2 = os.path.join(file_dir, "Hello.zip")
    with ZipFile(file_dir2) as zip_file:
        print(zip_file.namelist())
        text = zip_file.read('Hello.txt')
        print(text)
        zip_file.extract('Hello.txt', path="zip")